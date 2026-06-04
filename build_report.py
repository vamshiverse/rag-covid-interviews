"""Build the editable RAG experiment-tracker spreadsheet.

Each row = one pipeline configuration (the major design choices as columns) plus
the resulting evaluation scores. Seeded with the two real measured runs
(MMR off vs on) and a set of "experiment idea" rows to fill in as you tune the
pipeline. Re-run this anytime to regenerate; the CURRENT row is pulled live from
storage/eval_results.json.

    python build_report.py
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "RAG_Experiment_Tracker.xlsx"

# ---- columns: (header, width) ----
COLS = [
    ("Run ID", 14), ("Date", 12), ("Notes / what changed", 40),
    # --- design choices ---
    ("Chunking", 22), ("Embedding", 30), ("Retrieval", 26),
    ("MMR", 7), ("MMR λ", 7), ("top_k", 7), ("cand_k", 7), ("RRF k", 7),
    ("Metadata filter", 16), ("Generator", 14), ("Judge", 16), ("Fallback thr", 11),
    # --- metrics (0-1) ---
    ("Context Relevance", 11), ("Groundedness", 11), ("Answer Relevance", 11),
    ("RAG Triad", 10), ("Hit Rate@5", 10), ("MRR", 8), ("Context Recall", 11),
    ("Context Precision", 11), ("Answer Correctness", 12), ("Citation Grounding", 12),
    ("Fallback Correct", 11), ("Latency (ms)", 11),
]
# metric column indices (1-based)
P, Q, R, S = 16, 17, 18, 19          # ctx_rel, ground, ans_rel, triad(formula)
METRIC_FIRST, METRIC_LAST = 16, 26   # Context Relevance .. Fallback Correct (0-1)
LAT_COL = 27

DARK = "0D3B66"; MID = "1D6FB8"; LIGHT = "EEF4FB"
thin = Side(style="thin", color="D9DEE7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _load_current():
    p = ROOT / "storage" / "eval_results.json"
    if not p.exists():
        return None
    m = json.load(open(p, encoding="utf-8"))["aggregate"]["all_metrics"]
    return m


def metrics_row(m):
    """map an all_metrics dict to the metric cell values (ctx,ground,ans,_,hit,mrr,rec,prec,corr,cite,fb,lat)"""
    return [round(m["context_relevance"], 3), round(m["groundedness"], 3),
            round(m["answer_relevance"], 3), None,  # triad = formula
            round(m["hit_rate"], 3), round(m["mrr"], 3), round(m["context_recall"], 3),
            round(m["context_precision"], 3), round(m["answer_correctness"], 3),
            round(m["citation_grounding"], 3), round(m["fallback_correct"], 3),
            round(m["latency_ms_total"])]


# Baseline (MMR OFF) — measured before MMR was added.
BASELINE_METRICS = [0.493, 0.726, 0.671, None, 0.933, 0.839, 0.889, 0.267,
                    0.628, 1.0, 0.889, 188]

CONFIG_BASE = dict(chunk="Q&A-aware (Q+A together)",
                   embed="sentence-transformers / MiniLM-L6",
                   retr="Hybrid (dense + BM25 → RRF)",
                   topk=5, candk=20, rrfk=60, mfilter="None",
                   gen="extractive", judge="math (embedding)", fbthr=0.40)


def main():
    cur = _load_current()
    cur_metrics = metrics_row(cur) if cur else BASELINE_METRICS

    wb = Workbook()
    ws = wb.active
    ws.title = "Experiments"

    # header
    for j, (h, w) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 34

    def config_cells(cfg, mmr, lam):
        return [cfg["chunk"], cfg["embed"], cfg["retr"], mmr, lam, cfg["topk"],
                cfg["candk"], cfg["rrfk"], cfg["mfilter"], cfg["gen"], cfg["judge"], cfg["fbthr"]]

    # ---- seeded real rows ----
    rows = []
    rows.append(["R01-baseline", "2026-06-04", "Pure hybrid, NO MMR (the 'before')"]
                + config_cells(CONFIG_BASE, "Off", None) + BASELINE_METRICS)
    rows.append(["R02-current", "2026-06-04", "DEPLOYED: hybrid + MMR + metadata filtering available"]
                + config_cells(CONFIG_BASE, "On", 0.8) + cur_metrics)

    # ---- experiment-idea rows (config filled, metrics blank to fill after running) ----
    def idea(rid, note, **over):
        cfg = dict(CONFIG_BASE); cfg.update(over)
        mmr = over.pop("mmr", "On"); lam = over.pop("lam", 0.8)
        return [rid, "", note] + config_cells(cfg, over.get("_mmr", mmr), over.get("_lam", lam)) + [None]*12

    idea_rows = [
        idea("R03", "TF-IDF embeddings (no PyTorch) — lighter; measure quality drop",
             embed="tfidf (1-2gram)", fbthr=0.06),
        idea("R04", "top_k = 3 — tighter context, may raise Context Relevance", topk=3),
        idea("R05", "MMR λ = 0.5 — more diversity; watch Context Recall", _mmr="On", _lam=0.5),
        idea("R06", "Dense-only retrieval (drop BM25)", retr="Dense only"),
        idea("R07", "BM25-only retrieval (drop embeddings)", retr="BM25 only"),
        idea("R08", "LLM generation + LLM-as-judge (OpenAI)", gen="openai gpt-4o-mini", judge="LLM-as-judge"),
        idea("R09", "Context Relevance = mean of top-3 (fairer metric)"),
    ]

    all_rows = rows + idea_rows
    for i, rowvals in enumerate(all_rows, start=2):
        for j, val in enumerate(rowvals, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = BORDER
            if j == S:  # RAG Triad formula (mean of the 3 triad metrics)
                c.value = f"=IFERROR(AVERAGE({get_column_letter(P)}{i}:{get_column_letter(R)}{i}),\"\")"
            if METRIC_FIRST <= j <= METRIC_LAST:
                c.number_format = "0.000"
                c.alignment = Alignment(horizontal="center")
            elif j == LAT_COL:
                c.number_format = "0"
                c.alignment = Alignment(horizontal="center")
            elif j in (7, 8, 9, 10, 11, 15):
                c.alignment = Alignment(horizontal="center")
        # highlight the deployed row
        if rowvals[0] == "R02-current":
            for j in range(1, len(COLS) + 1):
                ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=LIGHT)
        # zebra-stripe the idea rows lightly
        if i > 3:
            ws.cell(row=i, column=1).font = Font(name="Arial", size=10, italic=True, color="7A8493")

    last = len(all_rows) + 1
    # freeze header + first 3 id/notes columns
    ws.freeze_panes = "D2"

    # conditional formatting: green=high for 0-1 metrics, reversed for latency
    rng = f"{get_column_letter(METRIC_FIRST)}2:{get_column_letter(METRIC_LAST)}{last+10}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0, start_color="F4A582",
        mid_type="num", mid_value=0.6, mid_color="FFFFBF",
        end_type="num", end_value=1, end_color="A6D96A"))
    lat_rng = f"{get_column_letter(LAT_COL)}2:{get_column_letter(LAT_COL)}{last+10}"
    ws.conditional_formatting.add(lat_rng, ColorScaleRule(
        start_type="min", start_color="A6D96A", end_type="max", end_color="F4A582"))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ---------------- Column guide sheet ----------------
    g = wb.create_sheet("Column guide")
    g.column_dimensions["A"].width = 24
    g.column_dimensions["B"].width = 92
    g.cell(1, 1, "How to use this tracker").font = Font(name="Arial", bold=True, size=12, color=DARK)
    guide = [
        ("", ""),
        ("Workflow", "Change ONE part of the pipeline (an env var in rag_app/config.py, or the UI), run "
                     "`python run_eval.py`, then add a new row here with the config + the printed scores."),
        ("", ""),
        ("— Design choices —", ""),
        ("Chunking", "How documents are split. Current: Q&A-aware (each chunk = one question + its full answer)."),
        ("Embedding", "Vector model for dense retrieval. MiniLM (semantic) / tfidf (lexical) / openai."),
        ("Retrieval", "Hybrid (dense+BM25→RRF) / Dense only / BM25 only."),
        ("MMR / λ", "Maximal Marginal Relevance de-duplication. λ=1 pure relevance, λ=0 pure diversity."),
        ("top_k", "Chunks passed to the answerer. cand_k = candidates pulled per signal. RRF k = fusion constant."),
        ("Metadata filter", "Restrict to country/specialty/topic before ranking."),
        ("Generator", "extractive (no key) / openai / anthropic."),
        ("Judge", "How metrics are scored: math (embedding proxies, no key) / LLM-as-judge."),
        ("Fallback thr", "Min similarity to answer; below it the engine abstains with a reason."),
        ("", ""),
        ("— Metrics (0–1, higher better; RAG Triad = mean of first three) —", ""),
        ("Context Relevance", "Are retrieved chunks on-topic for the question? (RAG Triad)"),
        ("Groundedness", "Is every claim in the answer supported by the retrieved context? (RAG Triad)"),
        ("Answer Relevance", "Does the answer actually address the question? (RAG Triad)"),
        ("Hit Rate@5 / MRR", "Did we retrieve a gold chunk, and how high was the first one?"),
        ("Context Recall", "Fraction of expected gold chunks retrieved."),
        ("Context Precision", "Fraction of retrieved chunks that are gold. NOTE: low ceiling (~0.2–0.4) because "
                              "the golden set lists only 1–3 ideal chunks per question — not a quality failure."),
        ("Answer Correctness", "Similarity of the answer to the ideal/reference answer."),
        ("Citation Grounding", "Fraction of quoted citation lines that truly appear in their cited source chunk."),
        ("Fallback Correct", "Fraction of questions where the engine abstained exactly when it should."),
        ("Latency (ms)", "Per-question time. Machine-dependent; first call includes model warm-up — compare on the same machine."),
    ]
    for i, (a, b) in enumerate(guide, start=2):
        ca = g.cell(i, 1, a); cb = g.cell(i, 2, b)
        ca.font = Font(name="Arial", bold=a.startswith("—") or a in ("Workflow",), size=10,
                       color=DARK if a.startswith("—") else "000000")
        cb.font = Font(name="Arial", size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("Seeded rows: R01-baseline (MMR off), R02-current (deployed) + 7 experiment-idea rows.")


if __name__ == "__main__":
    main()
