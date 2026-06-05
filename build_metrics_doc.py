"""Generate Metrics_Explained.xlsx — one row per metric with intuition, source
library, and the exact mathematical formula (as actually implemented)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK = "0D3B66"; MID = "1D6FB8"; BAND = "EEF4FB"
thin = Side(style="thin", color="D9DEE7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("Metric", 22), ("Stage", 12), ("Intuition — what it measures", 52),
        ("Source / library", 44), ("Exact formula  (default 'math' mode)", 60),
        ("Range · good", 18)]

# (metric, stage, intuition, source, formula, range)
ROWS = [
 ("Context Relevance", "Retrieve (Triad)",
  "Are the retrieved chunks actually on-topic for the question? Penalises pulling "
  "irrelevant context.",
  "Concept: RAG Triad (TruLens / 'Building & Evaluating Advanced RAG' course). "
  "Impl: custom — cosine on sentence-transformers (MiniLM) via numpy. LLM mode: LLM-as-judge.",
  "CR = (1/k)·Σ_{i=1..k} max(0, cos(q, cᵢ))   — mean question–chunk cosine over the k "
  "retrieved chunks.   [LLM mode: judge_score ÷ 10]",
  "0–1 · higher"),

 ("Groundedness / Faithfulness", "Ground (Triad)",
  "Is every claim in the answer actually supported by the retrieved context (no made-up facts)?",
  "Concept: RAG Triad (TruLens / course). Impl: custom sentence-level cosine (numpy + MiniLM). "
  "LLM mode: LLM-as-judge.",
  "G = (1/|Sₐ|)·Σ_{s∈Sₐ} max_{t∈S_ctx} cos(s, t)   — Sₐ = answer sentences, S_ctx = sentences "
  "of the used chunks; each answer sentence scored by its best support.",
  "0–1 · higher"),

 ("Answer Relevance", "Answer (Triad)",
  "Does the answer actually address what was asked (not off-topic / evasive)?",
  "Concept: RAG Triad (TruLens / course). Impl: custom cosine (numpy + MiniLM). LLM mode: judge.",
  "AR = max(0, cos(q, a))   — cosine between the question embedding q and the full answer "
  "embedding a.   [LLM mode: judge_score ÷ 10]",
  "0–1 · higher"),

 ("RAG Triad (overall)", "Triad",
  "Single headline number combining the three triad metrics.",
  "Aggregate of the three above (TruLens RAG Triad).",
  "Triad = (CR + G + AR) / 3",
  "0–1 · higher"),

 ("Hit Rate @k", "Retrieve",
  "Did we retrieve at least one of the expected 'gold' chunks within the top-k?",
  "Standard Information Retrieval metric. Impl: custom (numpy/Python). (Also used in "
  "LlamaIndex/LangChain retriever evals.)",
  "HR@k = 1 if (retrieved_top_k ∩ G) ≠ ∅, else 0   — per query; reported as the mean over the "
  "golden set.  G = set of expected chunk IDs, k = top_k (=5).",
  "0–1 · higher"),

 ("MRR (Mean Reciprocal Rank)", "Retrieve",
  "How high up was the FIRST correct chunk? Rewards putting gold near the top.",
  "Standard IR metric. Impl: custom (numpy/Python).",
  "RR = 1 / rank*,  rank* = position of the first retrieved chunk that is in G (RR=0 if none). "
  "MRR = mean(RR) over all queries.",
  "0–1 · higher"),

 ("Context Precision @k", "Retrieve",
  "Of the chunks we fed the answerer, what fraction were truly relevant (signal vs noise)?",
  "Standard IR metric. Impl: custom.",
  "P@k = |retrieved_top_k ∩ G| / k.   NOTE: low ceiling here (~0.2–0.4) because the golden set "
  "lists only 1–3 ideal chunks per question — not a quality failure.",
  "0–1 · higher*"),

 ("Context Recall @k", "Retrieve",
  "Of all the expected gold chunks, what fraction did we manage to retrieve?",
  "Standard IR metric. Impl: custom.",
  "R@k = |retrieved_top_k ∩ G| / |G|",
  "0–1 · higher"),

 ("Answer Correctness", "Answer",
  "How close is the generated answer to the curated 'ideal' answer in meaning?",
  "Custom — semantic similarity (cosine) to the reference answer; concept akin to RAGAS "
  "'answer similarity/correctness'. LLM mode: LLM-as-judge vs reference.",
  "AC = max(0, cos(a, a*))   — a* = ideal/reference answer from the golden set.   "
  "[LLM mode: judge_score ÷ 10]",
  "0–1 · higher"),

 ("Citation Grounding", "Ground",
  "Do the quoted citation lines REALLY appear in the source chunk they cite (no fabricated quotes)?",
  "Custom — token-overlap (containment) check. Directly verifies the grounding requirement.",
  "CG = (1/|C|)·Σ_{c∈C} 1[ overlap(quoteᶜ, srcᶜ) ≥ 0.8 ],  where "
  "overlap(x,y) = |tok(x) ∩ tok(y)| / |tok(x)|.  C = citations returned.",
  "0–1 · higher (1.0 ideal)"),

 ("Citation Source Precision", "Ground",
  "Of the sources we cited, what fraction were among the expected gold sources?",
  "Custom — set precision of cited chunk IDs vs gold.",
  "CSP = |cited_ids ∩ G| / |cited_ids|",
  "0–1 · higher"),

 ("Citation Source Recall", "Ground",
  "Of the expected gold sources, what fraction did we cite?",
  "Custom — set recall of cited chunk IDs vs gold.",
  "CSR = |cited_ids ∩ G| / |G|",
  "0–1 · higher"),

 ("Fallback Correctness", "Ops",
  "Did the engine abstain (say 'I can't answer') exactly when it should — and answer when it should?",
  "Custom — abstention accuracy (our anti-hallucination design).",
  "FC = (1/N)·Σ 1[ is_fallback(item) == (NOT answerable(item)) ]   over all N golden items.",
  "0–1 · higher"),

 ("Latency (ms)", "Ops",
  "Wall-clock time to answer one question (retrieval + generation).",
  "Custom — Python time.perf_counter().",
  "L = t_total = t_retrieval + t_generation (milliseconds). Machine-dependent; first call "
  "includes model warm-up.",
  "ms · lower"),
]

wb = Workbook(); ws = wb.active; ws.title = "Metrics"

ws.merge_cells("A1:F1")
t = ws["A1"]; t.value = "RAG Evaluation Metrics — definitions, sources & formulas"
t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=DARK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:F2")
n = ws["A2"]
n.value = ("Notation:  q = question,  a = answer,  a* = ideal answer,  cᵢ = i-th retrieved chunk,  "
           "G = set of expected 'gold' chunk IDs,  k = top_k (=5).  Embeddings are L2-normalised, "
           "so cos(a,b) = a·b (dot product), computed with numpy on sentence-transformers "
           "all-MiniLM-L6-v2.  'LLM mode' = optional LLM-as-judge (TruLens-style feedback functions).")
n.font = Font(name="Arial", size=9, italic=True, color="44515F")
n.fill = PatternFill("solid", fgColor="F4F7FB")
n.alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[2].height = 46

hdr = 3
for j, (h, w) in enumerate(COLS, start=1):
    c = ws.cell(hdr, j, h)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=MID)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[hdr].height = 30

for i, row in enumerate(ROWS, start=hdr + 1):
    band = (i - hdr) % 2 == 0
    for j, val in enumerate(row, start=1):
        c = ws.cell(i, j, val)
        c.font = Font(name="Arial", size=10, bold=(j == 1),
                      color=DARK if j == 1 else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal="center" if j in (2, 6) else "left")
        c.border = BORDER
        if band:
            c.fill = PatternFill("solid", fgColor=BAND)

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A{hdr}:F{hdr}"

wb.save("Metrics_Explained.xlsx")
print("Wrote Metrics_Explained.xlsx with", len(ROWS), "metrics")
